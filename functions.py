"""
Help functions for Finance Manager 3.0
"""
import json
import logging
import os.path
import sqlite3
import datetime
import urllib.request
# import sys
# import dropbox
# import requests
# import base64

import qtawesome as qta

# from dropbox.files import WriteMode
# from dropbox.exceptions import ApiError, AuthError
from PyQt6.QtWidgets import QMessageBox
from PyQt6.QtCore import QDate
from openpyxl import Workbook, load_workbook
from os.path import isfile


# Popup dialog
def popup_message(text: str, title: str = "Upozorenje", icon=QMessageBox.Icon.Information, style="") -> QMessageBox:
    """Creates warning popup box"""

    mb = QMessageBox()
    mb.setWindowTitle(title)
    mb.setIcon(icon)
    mb.setStandardButtons(QMessageBox.StandardButton.Ok)
    mb.setStyleSheet(style)
    mb.setText(text)

    return mb


# Set table layout
def set_table_layout(table):
    """Set table header height and width, alignment"""

    # Set column width
    table.setColumnWidth(0, 150)
    table.setColumnWidth(1, 100)
    table.setColumnWidth(2, 465)
    table.setColumnWidth(3, 150)
    table.setColumnWidth(4, 150)
    table.setColumnWidth(5, 300)

    # Set row height
    for row in range(table.rowCount()):
        table.setRowHeight(row, 30)


# Set Tree layout - income/outcome
def set_tree_layout(tree):
    """Set treeview columns width - income/outcome"""

    tree.setColumnWidth(0, 45)
    tree.setColumnWidth(1, 325)
    tree.setColumnWidth(2, 275)
    tree.setColumnWidth(3, 100)
    tree.setColumnWidth(4, 100)
    tree.setColumnWidth(5, 250)


# Set Tree layout - reports
def set_report_tree_layout(tree):
    """Set treeview columns width - reports"""
    tree.setColumnWidth(0, 50)
    tree.setColumnWidth(1, 375)
    tree.setColumnWidth(2, 200)
    tree.setColumnWidth(3, 200)


# Load config file
def load_config(path="config.json"):
    """Load settings from config file (json)"""
    with open(path, 'r', encoding="utf-8") as config_file:
        config = json.load(config_file)
    return config


def load_logging_settings(path="settings.db"):
    """Load logging parameters"""
    log_settings = get_data_from_database(path, "SELECT * FROM settings WHERE setting = 'log_filename' "
                                                "OR setting = 'log_level' OR setting = 'log_format'")
    return [i[1] for i in log_settings]


# Database functions
def get_data_from_database(database_path, query):
    """
    Gets data from database.

    Keyword arguments:
    database_path: full or relative path to database
    query: string used in sqlite execute
    """

    if not os.path.isfile(database_path):
        print("File does not exists!!")
        return

    # Open database
    conn = sqlite3.connect(database_path)

    # Create cursor
    cursor = conn.cursor()

    # Execute query
    cursor.execute(query)

    # Save data
    data = cursor.fetchall()

    # Close database
    conn.close()

    # Return data
    return data


def update_database(database_path, query):
    """
    Updates and saves changes to database.

    Keyword arguments:
    database_path: full or relative path to database
    query: string used in sqlite execute
    """
    # Open database
    conn = sqlite3.connect(database_path)

    # Create cursor
    cursor = conn.cursor()

    # Execute query
    cursor.execute(query)

    # Save changes
    conn.commit()

    # Close database
    conn.close()


def save_data(database_path, query):
    """Save data to database"""

    conn = sqlite3.connect(database_path)
    c = conn.cursor()
    c.execute(query)
    conn.commit()
    conn.close()


def set_id(database_path, table):
    """Creates unique ID number for entry"""

    query = f"SELECT id FROM {table}"
    conn = sqlite3.connect(database_path)
    c = conn.cursor()
    # result = None
    id_number = 1
    try:
        c.execute(query)
        result = c.fetchall()
        n = len(result)
        ids = []
        for i in range(n):
            ids.append(result[i][0])

        while id_number in ids:
            id_number = id_number + 1

        return id_number

    except sqlite3.Error as e:
        print(f"The error '{e}' occurred")


def table_exists(database_path, table):
    """Check if table exists in given database - return TRUE or FALSE"""
    if os.path.isfile(database_path):
        # Create query
        query = f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table}'"

        # Check for table in database
        if len(get_data_from_database(database_path, query)) == 1:
            return True
        else:
            return False


def calculate_year_totals(year, database):
    """
    Calculate income/outcome yearly total
    :param year: int year
    :param database: string income/outcome
    :return: float total - sum of all income/outcome in database
    """
    income = get_data_from_database(database,
                                    f"SELECT amount FROM income WHERE day > '{year}-01-01' AND day < '{year}-12-31'")
    outcome = get_data_from_database(database,
                                     f"SELECT amount FROM outcome WHERE day > '{year}-01-01' AND day < '{year}-12-31'")

    return sum([i[0] for i in income]), sum([i[0] for i in outcome])


def calculate_daily_totals(database):
    """Calculate daily money traffic"""
    income = get_data_from_database(database,
                                    f"SELECT amount FROM income WHERE day = '{QDate.currentDate().toPyDate()}'")

    outcome = get_data_from_database(database,
                                     f"SELECT amount FROM outcome WHERE day = '{QDate.currentDate().toPyDate()}'")

    return sum(i[0] for i in income), sum([i[0] for i in outcome])


# Load data from bank report
def decode_input_file(path):
    """Decodes data from bank report"""
    # Učitava podatke iz izvoda i vraca listu s podacima
    data = []
    nlines = sum(1 for _ in open(path))
    with open(path, 'r') as file:
        for line in range(2, nlines):
            data.append(file.readlines(line))

    br_izvoda = data[1][0][166:169]
    yy_izvoda = data[0][0][72:76]

    data = data[2:(nlines - 3)]

    result = []
    for i in data:
        text = i[0]
        text = text.replace("Æ", "Ć")
        text = text.replace("È", "Č")
        text = text.replace("æ", "ć")
        text = text.replace("è", "č")
        text.upper()

        income_outcome_id = (text[0:2])
        # iban = text[2:36].strip()
        uplatitelj = text[36:106].strip()
        # adresa = text[106:141].strip()
        # mjesto = text[141:176].strip()
        datum = text[176:184].strip()
        iznos = float(text[228:242].strip()) / 100
        poziv_na_broj = text[268:294].strip()
        opis = text[298:480].strip()

        # PRIHOD/RASHOD
        if income_outcome_id == "10":
            i_o_value = "Rashod"
        elif income_outcome_id == "20":
            i_o_value = "Prihod"
        else:
            i_o_value = "Nepoznato"

        # DATUM
        yy = datum[0:4]
        mm = datum[4:6]
        dd = datum[6:8]
        date_value = "{}.{}.{}.".format(dd, mm, yy)

        result.append([i_o_value, opis, iznos, date_value, br_izvoda, uplatitelj, poziv_na_broj])

    return result, yy_izvoda


def write_to_excel(data, filename, sheet=None, header=None):
    """
    Writes data to Excel file
    :param data: list of lists
    :param filename: path of file to write
    :param sheet: sheetname
    :param header: column names
    :return: Nothing
    """

    # Create workbook
    if isfile(filename):
        wb = load_workbook(filename)
        if sheet in wb.get_sheet_names():
            ws = wb.get_sheet_by_name(sheet)
            wb.remove_sheet(ws)
        ws = wb.create_sheet(sheet)
    else:
        wb = Workbook()
        ws = wb.create_sheet(title=sheet)

    # Add header if given
    if header:
        ws.append(header)

    # Add data to sheet
    for row in data:
        ws.append(row)

    # Remove empty sheet
    if "Sheet" in wb.get_sheet_names():
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))

    # Save workbook
    wb.save(filename=filename)


def calculate_totals(start, stop, table, database):
    """
    Calculates income/outcome totals

    :param start: QDate, start date
    :param stop: QDate, end date
    :param table: income OR outcome
    :param database
    :return: dates, total
    """

    # Get dates
    query = f"SELECT day FROM {table} WHERE day > '{start}' AND day < '{stop}'"
    dates = get_data_from_database(database, query)
    dates = list(set([i[0] for i in dates]))
    dates.sort()

    # Calculate daily sums
    total = []
    for idate, date in enumerate(dates):
        data = get_data_from_database(database, f"SELECT amount FROM {table} WHERE day = '{date}'")
        total.append(sum([i[0] for i in data]))

    return dates, total


def handle_dates(dates):
    """Format dates for database - removes zeroes representing time"""

    for idate, date in enumerate(dates):
        if len(date) == 10:
            dates[idate] = datetime.datetime.strptime(date, "%Y-%m-%d")
        elif len(date) == 19:
            dates[idate] = datetime.datetime.strptime(date, "%Y-%m-%d 00:00:00")

    return dates


def create_stylesheet(style: str = "Default") -> str:
    """
    Creates stylesheet text
    :param style: Stylesheet
    :return: Stylesheet text
    """

    background = None
    text = None
    widget_background = None
    button_hover = None
    disabled_tab_text_color = None

    if style == 1 or style == "BlueGreyDark":
        background = "#102A43"
        text = "#F0F4F8"
        widget_background = "#243B53"
        button_hover = "#334E68"
        disabled_tab_text_color = "#778899"
    elif style == 2 or style == "BlueGreyLight":
        background = "#BCCCDC"
        text = "#102A43"
        widget_background = "#9FB3C8"
        button_hover = "#B3C5D7"
        disabled_tab_text_color = "#778899"
    elif style == 3 or style == "CoolGrey":
        background = "#1F2933"
        text = "#F5F7FA"
        widget_background = "#323F4B"
        button_hover = "#3E4C59"
        disabled_tab_text_color = "#778899"
    elif style == 4 or style == "DarkRed":
        background = "#1A0000"
        text = "#FF9999"
        widget_background = "#660000"
        button_hover = "#990000"
        disabled_tab_text_color = "#778899"
    elif style == 5 or style == "WarmGrey":
        background = "#27241D"
        text = "#FAF9F7"
        widget_background = "#423D33"
        button_hover = "#504A40"
        disabled_tab_text_color = "#778899"
    elif style == 6 or style == "Green":
        background = "#004A40"
        text = "#FAF9F7"
        widget_background = "#005146"
        button_hover = "#007666"
        disabled_tab_text_color = "#778899"
    else:
        style = "Default"

    # Create stylesheet
    if style not in ["Default", 0]:
        stylesheet_text = \
            f"""
                QWidget{{
                    background-color: {background};
                    color: {text};
                    font: 12pt "Calibri";
                }}
                
                .FinanceManager {{
                    width: 100%;
                }}
                
                #club {{
                    font: 16pt "Calibri";
                }}
                
                QPushButton, QComboBox, QDateEdit, QSpinBox, QLineEdit, QTabBar:tab, QHeaderView::section,
                QTreeWidget, QTableCornerButton::section {{
                    background-color: {widget_background};
                }}
                
                QCalendarWidget QWidget {{
                    alternate-background-color: {widget_background};
                }}
                
                QLabel {{
                    width: 100px;
                    height: 25px;
                }}
                
                QPushButton {{
                    width: 125px;
                    height: 30px;
                    max-width: 125px;
                    max-height:30px;
                    margin: 0px 5px;
                    border-radius: 10px;
                    color: {text};
                }}
                
                #add_row, #remove_row, #btn_hide, #btn_network, #btn_exit, #btn_backup {{
                    width: 40px;
                    height: 30px;
                }}
                
                QPushButton:hover {{
                    background-color: {button_hover};
                    border: 2px solid {widget_background};
                }}
                
                QTabBar:tab:selected {{
                    background-color: {button_hover};
                    border: 2px solid {widget_background};
                }}
                
                QDateEdit, QHeaderView::section, QComboBox {{
                    border: 1px solid {button_hover};
                }}
                
                QTabBar:tab {{
                    margin-right: 5px;
                    width: 30px;
                    height: 71px;
                    font: 10pt "Calibri";
                    font-weight: bold;
                }}
                
                QTabBar:tab:disabled {{
                    color: {disabled_tab_text_color}
                }}
                
                QTabWidget::pane {{
                    border: none;
                }}
                
                QHeaderView::section {{
                    font-size: 14px;
                }}
                
                QTreeWidget {{
                    alternate-background-color: {button_hover};
                }}
                
                QTableWidget {{
                    color: white;
                    gridline-color: {text};
                }}
                
                Line {{
                    background-color: {text};
                }}
            """

        # Set Icons defaults
        qta.set_defaults(color=text, scale_factor=0.70)

        return stylesheet_text

    else:
        stylesheet_text = """
        QWidget{
            font: 12pt "Calibri";
        }
        .FinanceManager {
            width: 100%;
        }
        #club {
            font: 16pt "Calibri";
        }
        QLabel {
            width: 100px;
            height: 25px;
        }
        QPushButton {
            width: 125px;
            height: 30px;
            max-width: 125px;
            max-height:30px;
            margin: 0px 5px;
        }
        #add_row, #remove_row, #btn_hide, #btn_network {
            width: 40px;
            height: 30px;
        }
        QTabBar:tab {
            margin-right: 5px;
            width: 30px;
            height: 71px;
            font: 10pt "Calibri";
            font-weight: bold;
        }
        
        QTabBar:tab:selected {
            border: 1px solid grey;
            border-right: 2px solid red;
            background-color: rgb(200, 200, 200)
        }
        
        QTabWidget::pane {
            border: none;
        }
        """

        return stylesheet_text


def create_link_style(style="Default"):
    """Returns CSS for link format"""
    if style == 1 or style == "BlueGreyDark":
        text = "#F0F4F8"
    elif style == 2 or style == "BlueGreyLight":
        text = "#102A43"
    elif style == 3 or style == "CoolGrey":
        text = "#F5F7FA"
    elif style == 4 or style == "DarkRed":
        text = "#FF9999"
    elif style == 5 or style == "WarmGrey":
        text = "#FAF9F7"
    elif style == 6 or style == "Green":
        text = "#FAF9F7"
    else:
        text = "black"

    return f'style="color: {text}; text-decoration: none;"'


def backup_file(localfile, dest, token):
    """Upload/backup database to dropbox"""
    pass
    """
    APP_KEY = "2cnixj5ux96n7ob"
    APP_SECRET = "l0w3oh6gmnzu9hw"

    with dropbox.Dropbox(token) as dbx:
        # Check that the access token is valid
        try:
            dbx.users_get_current_account()
        except AuthError:
            sys.exit("ERROR: Invalid access token; try re-generating an "
                     "access token from the app console on the web.")

        # Create a backup of the current settings file
        with open(localfile, 'rb') as f:
            # We use WriteMode=overwrite to make sure that the settings in the file
            # are changed on upload
            logging.info("Uploading " + localfile + " to Dropbox as " + dest + "...")
            try:
                dbx.files_upload(f.read(), dest, mode=WriteMode('overwrite'))
            except ApiError as err:
                # This checks for the specific error where a user doesn't have
                # enough Dropbox space quota to upload this file
                if (err.error.is_path() and
                        err.error.get_path().reason.is_insufficient_space()):
                    sys.exit("ERROR: Cannot back up; insufficient space.")
                elif err.user_message_text:
                    logging.error(err.user_message_text)
                    sys.exit()
                else:
                    logging.error(err)
                    sys.exit()

            return True
    """


def download_database(dropbox_file_path, local_file_path, token):
    """Download database from Dropbox"""
    """
    with dropbox.Dropbox(token) as dbx:
        # Check that the access token is valid
        try:
            dbx.users_get_current_account()
        except AuthError:
            sys.exit("ERROR: Invalid access token; try re-generating an "
                     "access token from the app console on the web.")

        try:
            with open(local_file_path, 'wb') as f:
                metadata, result = dbx.files_download(path=dropbox_file_path)
                f.write(result.content)
        except Exception as e:
            print('Error downloading file from Dropbox: ' + str(e))
    """
    pass


def check_internet_connection():
    """Check for internet connection"""
    try:
        urllib.request.urlopen(url="http://google.com")
        logging.info("Internet connected")
        return True
    except Exception as error:
        logging.info(f"No internet connection: {error}")
        return False
